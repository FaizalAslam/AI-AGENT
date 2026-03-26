# server.py
from flask import Flask, render_template, request, jsonify
import threading
import json
import time
import logging
import webbrowser
import traceback
import os
import re

# â”€â”€ Core modules â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from modules import system_core, ui, config

# â”€â”€ Office Agent (Project 2) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from utils.command_buffer    import CommandBuffer
from utils import command_map
from executor.excel_executor import ExcelExecutor
from executor.word_executor  import WordExecutor
from executor.ppt_executor   import PowerPointExecutor
from parser.command_parser   import parse_command
from listener.keyboard_listener  import KeyboardListener
from listener.clipboard_listener import ClipboardListener
try:
    from listener.voice_listener import VoiceListener
    VOICE_MODULE_AVAILABLE = True
except Exception:
    VoiceListener = None
    VOICE_MODULE_AVAILABLE = False

# â”€â”€ Optional modules (graceful fallback) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
try:
    from modules import ocr_utils
    OCR_AVAILABLE = True
except Exception as e:
    print(f"OCR unavailable: {e}")
    OCR_AVAILABLE = False

try:
    from modules import pdf_utils
    PDF_AVAILABLE = True
except Exception as e:
    print(f"PDF unavailable: {e}")
    PDF_AVAILABLE = False

try:
    from modules import pdf_reader
    READER_AVAILABLE = True
except Exception as e:
    print(f"PDF Reader unavailable: {e}")
    READER_AVAILABLE = False

try:
    from modules import gui_automation
    GUI_AVAILABLE = True
except Exception as e:
    print(f"GUI unavailable: {e}")
    GUI_AVAILABLE = False

try:
    from modules import pdf_editor
    PDF_EDITOR_AVAILABLE = True
except Exception as e:
    print(f"PDF Editor unavailable: {e}")
    PDF_EDITOR_AVAILABLE = False

try:
    import keyboard
    KEYBOARD_AVAILABLE = True
except ImportError:
    print("keyboard not found â€” pip install keyboard")
    KEYBOARD_AVAILABLE = False

# â”€â”€ Logging â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
logging.basicConfig(
    filename="agent.log",
    level=logging.INFO,
    format="%(asctime)s - %(message)s",
    datefmt="%H:%M:%S",
    filemode="w"
)
logging.getLogger("werkzeug").setLevel(logging.ERROR)

# â”€â”€ Flask app â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
app = Flask(__name__)

# â”€â”€ Shared state â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
last_ocr = {"text": "", "pending": False}

# â”€â”€ Office Agent setup â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
OFFICE_APPS = {"excel", "word", "powerpoint", "ppt"}
OFFICE_OUTPUTS = {
    "excel": "output.xlsx",
    "word": "output.docx",
    "powerpoint": "output.pptx",
    "ppt": "output.pptx",
}
_cmd_buf            = CommandBuffer()
_clipboard_listener = ClipboardListener(_cmd_buf)
_keyboard_listener  = KeyboardListener(_handle_global_command := None, _cmd_buf)
_voice_listener     = VoiceListener(_handle_global_command) if VOICE_MODULE_AVAILABLE else None
voice_state         = {"enabled": False}


# â”€â”€ Office Agent helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _safe_speak(text):
    try:
        ui.speak(text)
    except Exception:
        pass


def _extract_office_agent_command(raw_text):
    text = (raw_text or "").strip()
    match = re.match(r"^agent\s*:\s*(excel|word|powerpoint|ppt)\s*:\s*(.+)$", text, re.IGNORECASE)
    if not match:
        return None, None
    app_name = match.group(1).lower().strip()
    command_text = match.group(2).strip()
    return app_name, command_text


def _resolve_actions(app_name, command_text):
    actions = parse_command(app_name, command_text)
    if actions:
        command_map.save_actions(app_name, command_text, actions)
        return command_text, actions, "json-parser"
    
    return command_text, [], "json-parser"


def _run_office_actions(app_name, actions, file_path=None):
    app_name = (app_name or "").lower().strip()
    output_path = (file_path or "").strip() or OFFICE_OUTPUTS.get(app_name, "output.xlsx")
    output_path = os.path.abspath(output_path)
    executed = []
    failures = []
    opened = False

    if app_name == "excel":
        from openpyxl import Workbook, load_workbook
        wb = load_workbook(output_path) if os.path.exists(output_path) else Workbook()
        ws = wb.active
        executor = ExcelExecutor(wb, ws)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok:
                executed.append(action_name)
            else:
                failures.append(f"{action_name} failed")
        wb.save(output_path)
    elif app_name == "word":
        from docx import Document
        doc = Document(output_path) if os.path.exists(output_path) else Document()
        executor = WordExecutor(doc)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok:
                executed.append(action_name)
            else:
                failures.append(f"{action_name} failed")
        doc.save(output_path)
    elif app_name in ("powerpoint", "ppt"):
        from pptx import Presentation
        prs = Presentation(output_path) if os.path.exists(output_path) else Presentation()
        executor = PowerPointExecutor(prs)
        for action in actions:
            ok = bool(executor.run(action))
            action_name = action.get("action", "unknown")
            if ok:
                executed.append(action_name)
            else:
                failures.append(f"{action_name} failed")
        prs.save(output_path)
    else:
        failures.append(f"Unsupported app: {app_name}")

    if not failures and os.path.exists(output_path):
        try:
            opened = bool(system_core.open_path(output_path))
        except Exception:
            opened = False

    return {
        "ok_count": len(executed),
        "total": len(actions),
        "executed": executed,
        "failures": failures,
        "output_path": output_path,
        "opened": opened,
    }


def _handle_global_command(raw_text):
    """Handles system-wide agent: <app>: <command> triggers."""
    try:
        app_name, command = _extract_office_agent_command(raw_text)
        if app_name and command:
            if app_name == "ppt":
                app_name = "powerpoint"
            cache_key, actions, source = _resolve_actions(app_name, command)
            if not actions:
                logging.warning(f"No parser match for global office command: {app_name}: {command}")
                return
            summary = _run_office_actions(app_name, actions)
            if summary["failures"] and cache_key:
                command_map.remove_action(app_name, cache_key)
            logging.info(
                f"Global office [{source}] {app_name}: {command} -> "
                f"{summary['ok_count']}/{summary['total']} | {summary['output_path']}"
            )
            _safe_speak(f"Executed {summary['ok_count']} actions in {app_name}")
            return

        txt = (raw_text or "").strip()
        low = txt.lower()
        if low.startswith("agent "):
            sys_cmd = txt[len("agent "):].strip()
            sys_cmd = sys_cmd.replace("  ", " ").strip(" .,:;!?")
            if sys_cmd.startswith(("open ", "launch ", "start ", "run ", "boot ")):
                success, message = system_core.find_and_launch(sys_cmd)
                _safe_speak(
                    f"Opening {system_core.normalize_app_name(sys_cmd)}"
                    if success else f"Could not open {sys_cmd}"
                )
                logging.info(f"Voice system open [{sys_cmd}] => {success}: {message}")
            elif sys_cmd.startswith(("close ", "shut ", "exit ")):
                success, message = system_core.close_app(sys_cmd)
                _safe_speak(
                    f"Closing {system_core.normalize_app_name(sys_cmd)}"
                    if success else f"Could not close {sys_cmd}"
                )
                logging.info(f"Voice system close [{sys_cmd}] => {success}: {message}")
    except Exception as e:
        logging.error(f"Global command error: {e}\n{traceback.format_exc()}")


# Patch the keyboard listener callback now that function is defined
_keyboard_listener.on_command = _handle_global_command
if _voice_listener:
    _voice_listener.on_command = _handle_global_command


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  ROUTES
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route("/")
def index():
    return render_template("index.html")


# â”€â”€ SYSTEM COMMANDS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/execute", methods=["POST"])
def execute():
    try:
        data = request.json
        cmd  = data.get("command", "").lower().strip()
        logging.info(f"Command: {cmd}")

        if cmd.startswith(("close ", "shut ", "exit ")):
            app_name = cmd.replace("close ", "").replace("shut ", "").replace("exit ", "").strip()
            success, message = system_core.close_app(app_name)
            _safe_speak(f"Closing {app_name}" if success else f"Could not close {app_name}")
            return jsonify(status="success" if success else "fail", message=message)

        app_name = system_core.normalize_app_name(cmd)
        success, message = system_core.find_and_launch(app_name)
        if success:
            _safe_speak(f"Opening {app_name}")
            return jsonify(status="success", message=message)

        _safe_speak(f"I couldn't find {app_name}. Please select it manually.")
        path = ui.manual_selector()
        if path:
            norm_app = system_core.normalize_app_name(app_name)
            config.save_memory(norm_app, path, is_store_app=False)
            launched = system_core.open_path(path)
            if launched:
                _safe_speak("Path saved. Opening now.")
                return jsonify(status="success", message="Manual Selection Saved")
            return jsonify(status="fail", message="Saved path, but launch failed")

        return jsonify(status="fail", message="Cancelled")

    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# â”€â”€ OFFICE AGENT â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/office/execute", methods=["POST"])
def office_execute():
    try:
        return _office_execute_impl(request.json or {})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


def _office_execute_impl(data):
    app_name = (data.get("app") or "").lower().strip()
    command = (data.get("raw") or "").strip()
    full = (data.get("command") or "").strip()
    file_path = (data.get("file_path") or data.get("file") or "").strip()

    if not command and full:
        parsed_app, parsed_command = _extract_office_agent_command(full)
        if parsed_app and not app_name:
            app_name = (parsed_app or "").strip()
        if parsed_command:
            command = (parsed_command or "").strip()
        elif app_name:
            command = full

    if app_name == "ppt":
        app_name = "powerpoint"

    if app_name not in OFFICE_APPS or not command:
        return jsonify(status="fail", message="Missing/invalid app or command")

    cache_key, actions, source = _resolve_actions(app_name, command)
    if not actions:
        return jsonify(status="fail", message="No matching command found in JSON parser", source=source)

    summary = _run_office_actions(app_name, actions, file_path=file_path)
    if summary["failures"] and cache_key:
        command_map.remove_action(app_name, cache_key)
        return jsonify(
            status="fail",
            message=f"✅ {summary['ok_count']}/{summary['total']} done | ❌ {' | '.join(summary['failures'])}",
            source=source,
            output_file=summary["output_path"]
        )

    return jsonify(
        status="success",
        message=f"✅ Executed {summary['ok_count']} actions. Output: {summary['output_path']}",
        source=source,
        output_file=summary["output_path"],
        opened=summary.get("opened", False)
    )


@app.route("/command", methods=["POST"])
def office_command():
    try:
        return _office_execute_impl(request.json or {})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# â”€â”€ OCR â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# VOICE CONTROL

@app.route("/voice/status", methods=["GET"])
def voice_status():
    if not _voice_listener:
        return jsonify(
            status="fail",
            available=False,
            enabled=False,
            message="Voice module unavailable. Install SpeechRecognition + PyAudio."
        )
    heard = _voice_listener.last_heard
    heard_age = time.time() - (_voice_listener.last_heard_at or 0)
    if heard_age > 8:
        heard = ""
    return jsonify(
        status="success",
        available=_voice_listener.available,
        enabled=_voice_listener.is_running,
        armed=_voice_listener.armed,
        armed_seconds=round(_voice_listener.armed_seconds_left, 1),
        heard=heard,
        error=_voice_listener.last_error
    )


@app.route("/voice/start", methods=["POST"])
def voice_start():
    if not _voice_listener:
        return jsonify(status="fail", message="Voice module unavailable")
    ok = _voice_listener.start()
    voice_state["enabled"] = bool(ok)
    return jsonify(
        status="success" if ok else "fail",
        message="Voice listener started" if ok else (_voice_listener.last_error or "Could not start voice listener")
    )


@app.route("/voice/stop", methods=["POST"])
def voice_stop():
    if not _voice_listener:
        return jsonify(status="fail", message="Voice module unavailable")
    _voice_listener.stop()
    voice_state["enabled"] = False
    return jsonify(status="success", message="Voice listener stopped")


@app.route("/ocr/snip", methods=["POST"])
def ocr_snip():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        ocr_utils.snip_queue.put("snip")
        try:
            path = ocr_utils.result_queue.get(timeout=60)
        except Exception:
            return jsonify(status="fail", message="Snip timed out")
        if not path:
            return jsonify(status="fail", message="Snip cancelled")
        text = ocr_utils.image_to_text(path)
        last_ocr["text"]    = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/screenshot", methods=["POST"])
def ocr_screenshot():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        path = ocr_utils.capture_fullscreen()
        text = ocr_utils.image_to_text(path)
        last_ocr["text"]    = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/file", methods=["POST"])
def ocr_file():
    try:
        if not OCR_AVAILABLE:
            return jsonify(status="fail", message="OCR not available")
        path = ui.file_selector(
            "Select an Image File",
            [("Images", "*.png *.jpg *.jpeg *.bmp *.tiff"), ("All Files", "*.*")]
        )
        if not path:
            return jsonify(status="fail", message="No file selected")
        text = ocr_utils.image_to_text(path)
        last_ocr["text"]    = text
        last_ocr["pending"] = False
        return jsonify(status="success", text=text, message=f"OCR complete â€” {len(text)} chars")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/read", methods=["POST"])
def ocr_read():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        threading.Thread(target=ocr_utils.speak_text, args=(text,), daemon=True).start()
        return jsonify(status="success", message="Speaking...")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/stop_read", methods=["POST"])
def ocr_stop_read():
    try:
        ocr_utils.stop_speaking()
        return jsonify(status="success", message="Stopped")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/poll", methods=["GET"])
def ocr_poll():
    try:
        if last_ocr.get("pending"):
            last_ocr["pending"] = False
            return jsonify(
                status="ready",
                text=last_ocr["text"],
                message=f"Hotkey OCR complete â€” {len(last_ocr['text'])} chars"
            )
        return jsonify(status="waiting")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/ocr/save_txt", methods=["POST"])
def ocr_save_txt():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        path = ocr_utils.save_as_txt(text)
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/save_pdf", methods=["POST"])
def ocr_save_pdf():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install fpdf2: pip install fpdf2")
        path = pdf_utils.create_report(text, title="OCR Result")
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/ocr/clipboard", methods=["POST"])
def ocr_clipboard():
    try:
        text = last_ocr.get("text", "")
        if not text:
            return jsonify(status="fail", message="No OCR text. Run OCR first.")
        ocr_utils.copy_to_clipboard(text)
        return jsonify(status="success", message="Copied to clipboard")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# â”€â”€ PDF READER â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/reader/open", methods=["POST"])
def reader_open():
    try:
        if not READER_AVAILABLE:
            return jsonify(status="fail", message="PDF reader module not found")
        path = ui.file_selector("Select PDF to Read", [("PDFs", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        threading.Thread(target=pdf_reader.start_reading, args=(path, 0), daemon=True).start()
        time.sleep(0.5)
        return jsonify(status="success", message="Reading started", **pdf_reader.get_status())
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/reader/pause", methods=["POST"])
def reader_pause():
    try:
        pdf_reader.pause_reading()
        return jsonify(status="success", message="Paused", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/resume", methods=["POST"])
def reader_resume():
    try:
        pdf_reader.resume_reading()
        return jsonify(status="success", message="Resumed", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/stop", methods=["POST"])
def reader_stop():
    try:
        pdf_reader.stop_reading()
        return jsonify(status="success", message="Stopped")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/next", methods=["POST"])
def reader_next():
    try:
        pdf_reader.next_page()
        return jsonify(status="success", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/prev", methods=["POST"])
def reader_prev():
    try:
        pdf_reader.prev_page()
        return jsonify(status="success", **pdf_reader.get_status())
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/speed", methods=["POST"])
def reader_speed():
    try:
        data = request.json
        pdf_reader.set_speed(data.get("speed", 150))
        return jsonify(status="success", message=f"Speed: {data.get('speed')} WPM")
    except Exception as e:
        return jsonify(status="fail", message=str(e))


@app.route("/reader/status", methods=["GET"])
def reader_status():
    try:
        return jsonify(pdf_reader.get_status())
    except Exception:
        return jsonify(
            is_reading=False, is_paused=False,
            current_page=0, total_pages=0, speed=150
        )


# â”€â”€ PDF TOOLS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/pdf/merge", methods=["POST"])
def pdf_merge():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install pypdf: pip install pypdf")
        paths = pdf_utils.ask(
            kind="openmultiple",
            title="Select PDFs to Merge (hold Ctrl for multiple)",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if not paths:
            return jsonify(status="fail", message="No files selected.")
        out = pdf_utils.merge_pdfs(paths)
        if not out:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"Merged {len(paths)} PDFs â†’ {out}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/pdf/split", methods=["POST"])
def pdf_split():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install pypdf: pip install pypdf")
        path = ui.file_selector("Select PDF to Split", [("PDFs", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected.")
        pages = pdf_utils.split_pdf(path)
        if not pages:
            return jsonify(status="fail", message="Save cancelled or no pages.")
        return jsonify(status="success", message=f"Split into {len(pages)} files")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/pdf/create", methods=["POST"])
def pdf_create():
    try:
        if not PDF_AVAILABLE:
            return jsonify(status="fail", message="Install fpdf2: pip install fpdf2")
        data  = request.json
        text  = data.get("text", "").strip()
        title = (data.get("title", "Report") or "Report").strip()
        if not text:
            return jsonify(status="fail", message="No text provided")
        path = pdf_utils.create_report(text, title=title)
        if not path:
            return jsonify(status="fail", message="Save cancelled.")
        return jsonify(status="success", message=f"PDF saved: {path}")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


# â”€â”€ PDF EDITOR â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/editor/open", methods=["POST"])
def editor_open():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        path = ui.file_selector("Select PDF to Edit", [("PDF Files", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        data = pdf_editor.extract_pdf_text(path)
        if data.get("status") != "success":
            return jsonify(status="fail", message=data.get("message", "Failed to open PDF"))
        return jsonify(
            status="success",
            file_path=path,
            pages=data["pages"],
            total_pages=data["total_pages"]
        )
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/render-page", methods=["POST"])
def editor_render_page():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data     = request.json
        pdf_path = data.get("file_path")
        page_num = data.get("page_num", 0)
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.render_page_as_image(pdf_path, page_num)
        if result.get("status") != "success":
            return jsonify(status="fail", message=result.get("message", "Render failed"))
        return jsonify(status="success", **{k: v for k, v in result.items() if k != "status"})
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/save", methods=["POST"])
def editor_save():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data     = request.json
        pdf_path = data.get("file_path")
        edits    = data.get("edits", [])
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.save_edited_pdf(pdf_path, edits)
        if result.get("status") != "success":
            return jsonify(status="fail", message=result.get("message", "Save failed"))
        return jsonify(status="success", message=result.get("message", "Saved successfully"))
    except Exception as e:
        logging.error(traceback.format_exc())
        return jsonify(status="fail", message=str(e))


@app.route("/editor/detect-form", methods=["POST"])
def editor_detect_form():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        path = ui.file_selector("Select PDF", [("PDF Files", "*.pdf")])
        if not path:
            return jsonify(status="fail", message="No file selected")
        fields = pdf_editor.detect_form_fields(path)
        return jsonify(
            status="success",
            is_form=len(fields) > 0,
            field_count=len(fields),
            fields=list(fields.keys()),
            file_path=path
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/editor/fill-form", methods=["POST"])
def editor_fill_form():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data      = request.json
        pdf_path  = data.get("file_path")
        form_data = data.get("form_data", {})
        if not pdf_path:
            return jsonify(status="fail", message="No file path provided")
        result = pdf_editor.fill_form(pdf_path, form_data)
        if result:
            return jsonify(status="success", message="Form saved successfully")
        return jsonify(status="fail", message="Save cancelled")
    except Exception as e:
        print(traceback.format_exc())
        return jsonify(status="fail", message=f"Error: {str(e)}")


@app.route("/editor/get-field-options", methods=["POST"])
def editor_get_field_options():
    try:
        if not PDF_EDITOR_AVAILABLE:
            return jsonify(status="fail", message="PDF Editor not available")
        data       = request.json
        pdf_path   = data.get("file_path")
        field_name = data.get("field_name")
        options    = pdf_editor.get_form_field_options(pdf_path, field_name)
        return jsonify(status="success", field_name=field_name, options=options)
    except Exception as e:
        return jsonify(status="fail", message=f"Error: {str(e)}")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  ENTRY POINT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

if __name__ == "__main__":

    # â”€â”€ OCR snip overlay (must be on main thread) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if OCR_AVAILABLE:
        threading.Thread(
            target=ocr_utils.run_snip_overlay_main_thread,
            daemon=True
        ).start()

    # â”€â”€ OCR keyboard hotkeys â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if KEYBOARD_AVAILABLE and OCR_AVAILABLE:
        keyboard.add_hotkey(
            "ctrl+shift+s",
            lambda: threading.Thread(
                target=ocr_utils.trigger_snip_and_ocr,
                args=(last_ocr,), daemon=True
            ).start()
        )
        keyboard.add_hotkey(
            "ctrl+shift+f",
            lambda: threading.Thread(
                target=ocr_utils.trigger_screenshot_and_ocr,
                args=(last_ocr,), daemon=True
            ).start()
        )
        print("ðŸ”‘  Ctrl+Shift+S â†’ Snip OCR  |  Ctrl+Shift+F â†’ Fullscreen OCR")

    # â”€â”€ Global Office Agent listeners â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    threading.Thread(
        target=_clipboard_listener.start,
        daemon=True, name="ClipboardListener"
    ).start()
    threading.Thread(
        target=_keyboard_listener.start,
        daemon=True, name="KeyboardListener"
    ).start()
    print("âŒ¨ï¸   Global agent listener active")
    print("     Type  agent: excel: <command>  anywhere + Enter")

    if _voice_listener and _voice_listener.available:
        if _voice_listener.start():
            voice_state["enabled"] = True
            print("Voice wake listener active (say: agent <app> <command>)")
        else:
            print(f"Voice listener not started: {_voice_listener.last_error}")

    # â”€â”€ Start Flask â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    flask_thread = threading.Thread(
        target=lambda: app.run(host="127.0.0.1", port=5000, debug=False),
        daemon=True
    )
    flask_thread.start()
    time.sleep(1)

    # â”€â”€ Open browser â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    webbrowser.open("http://127.0.0.1:5000")
    print("âœ…  Agent running at http://127.0.0.1:5000")

    # â”€â”€ Dialog listener MUST be on main thread â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if PDF_AVAILABLE:
        pdf_utils.run_dialog_listener()
    else:
        # Keep main thread alive if PDF not available
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\nðŸ‘‹ Agent stopped.")
